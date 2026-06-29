# -*- coding: utf-8 -*-
"""万能办公助手 — ExcelTools（合并·拆分·筛选·统计·导出·图表·VLOOKUP·条件格式）"""
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
from pathlib import Path
import os, sys, json, time, shutil, hashlib, threading
from datetime import datetime
from collections import Counter

from utils import OPENPYL_AVAILABLE, _OPENPYL_ERROR, safe_str, safe_cond_check

import openpyxl
from openpyxl.styles import PatternFill


class ExcelToolsMixin:
    """ExcelTools — 所有方法通过self访问OfficeAssistant的属性"""

    def _show_excel_tools(self):
        self.clear_content()
        self._section_header("Excel处理工具", "合并 · 拆分 · 筛选 · 统计 · 导出 · 图表 · VLOOKUP · 条件格式")
        self._show_tips(
            "点击下方卡片 → 选择Excel文件(.xlsx) → 自动处理 → 保存结果",
            "支持 .xlsx 格式，旧版 .xls 请先用Excel另存为 .xlsx"
        )
        if not OPENPYL_AVAILABLE:
            tk.Label(self.content_frame, text=f"⚠️ openpyxl未安装\n{_OPENPYL_ERROR}", fg="red",
                    font=("微软雅黑", 12), bg=self.colors['light']).pack(pady=60)
            return
        for rdata in [
            [("📊 合并Excel", "多文件纵向合并", self._merge_excel_dlg),
             ("✂ 拆分Excel", "按行数拆分为多个", self._split_excel_dlg),
             ("🔍 筛选数据", "按条件筛选/导出", self._filter_excel_dlg)],
            [("📈 数据统计", "最值/平均/计数/标准差", self._stats_excel_dlg),
             ("📊 导出CSV", "Excel→CSV (UTF-8)", self._excel_to_csv_dlg),
             ("📉 创建图表", "柱状图/折线图/饼图", self._excel_chart_dlg)],
            [("🔗 VLOOKUP", "跨表格匹配数据", self._vlookup_dlg),
             ("🎨 条件格式", "条件高亮/标记", self._conditional_format_dlg)],
        ]:
            row = tk.Frame(self.content_frame, bg=self.colors['light']); row.pack(fill=tk.X, padx=10)
            for title, desc, cb in rdata:
                self._create_card(row, title, desc, cb)

    def _merge_excel_dlg(self):
        files = filedialog.askopenfilenames(title="选择Excel文件(按顺序合并)", filetypes=[("Excel","*.xlsx")])
        if not files: return
        save = filedialog.asksaveasfilename(title="保存合并结果", defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not save: return
        def _go():
            try:
                wb_out = openpyxl.Workbook(); ws_out = wb_out.active; first = True
                for fp in files:
                    wb = openpyxl.load_workbook(fp, read_only=True); ws = wb.active
                    rows = list(ws.iter_rows(values_only=True)); wb.close()
                    if first:
                        ws_out.append([str(c) if c is not None else "" for c in rows[0]])
                        first = False
                    for r in rows[1:]:
                        ws_out.append([str(c) if c is not None else "" for c in r])
                wb_out.save(save); wb_out.close()
                self.root.after(0, lambda: self._show_success_dialog("完成", f"合并成功\n{len(files)}个文件\n{save}"))
            except Exception as e: self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
        self._run_thread(_go, done_msg="合并完成")

    def _split_excel_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        out_dir = filedialog.askdirectory(title="输出目录")
        if not out_dir: return
        win = tk.Toplevel(self.root); win.title("拆分设置"); win.geometry("350x150")
        win.transient(self.root); win.grab_set()
        tk.Label(win, text="每多少个行一个文件:", font=("微软雅黑",10)).pack(pady=10)
        rows_per = tk.IntVar(value=1000)
        tk.Spinbox(win, from_=1, to=100000, textvariable=rows_per, width=10).pack()
        def _work():
            win.destroy()
            try:
                wb = openpyxl.load_workbook(fp, read_only=True); all_rows = list(wb.active.iter_rows(values_only=True)); wb.close()
                if not all_rows: return
                hdr, src = all_rows[0], all_rows[1:]; stem = Path(fp).stem; i, op = 0, 0
                while i < len(src):
                    chunk = [hdr] + src[i:i+rows_per.get()]; i += rows_per.get(); op += 1
                    wb2 = openpyxl.Workbook(); ws2 = wb2.active
                    for r in chunk: ws2.append([str(c) if c is not None else "" for c in r])
                    wb2.save(str(Path(out_dir) / f"{stem}_part{op}.xlsx")); wb2.close()
                self.root.after(0, lambda: self._show_success_dialog("完成", f"拆分为 {op} 个文件\n{out_dir}"))
            except Exception as e: self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
        tk.Button(win, text="开始拆分", command=_work, bg=self.colors['primary'], fg="white",
                 font=("微软雅黑",11,"bold"), cursor="hand2", width=12).pack(pady=10)

    def _filter_excel_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        win = tk.Toplevel(self.root); win.title("筛选数据"); win.geometry("600x450")
        win.transient(self.root); win.grab_set()
        wb = openpyxl.load_workbook(fp, read_only=True); all_rows = list(wb.active.iter_rows(values_only=True)); wb.close()
        hdr = [str(c) for c in all_rows[0]] if all_rows else []
        log = scrolledtext.ScrolledText(win, height=12, font=("Consolas",9)); log.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        lb = tk.Listbox(win, height=4, font=("微软雅黑",9)); lb.pack(fill=tk.X, padx=10)
        for h in hdr: lb.insert(tk.END, h)
        e = tk.Entry(win, font=("微软雅黑",10)); e.pack(fill=tk.X, padx=10, pady=5)
        e.insert(0, ">0"); lb.selection_set(0)
        def _go():
            try:
                ci = lb.curselection(); ci = ci[0] if ci else 0
                cond = e.get().strip(); fltd = [all_rows[0]]
                ve = tk.StringVar(); h = hdr[ci]
                for r in all_rows[1:]:
                    v = r[ci]; val = float(v) if v is not None and str(v).replace('.','').replace('-','').isdigit() else 0
                    if safe_cond_check(val, cond): fltd.append(r)
                save = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
                if save:
                    wb2 = openpyxl.Workbook(); ws2 = wb2.active
                    for r in fltd: ws2.append([str(c) if c is not None else "" for c in r])
                    wb2.save(save); wb2.close()
                log.delete("1.0", tk.END)
                log.insert(tk.END, f"原始: {len(all_rows)-1} 行 | 筛选后: {len(fltd)-1} 行\n")
                for r in fltd[:30]: log.insert(tk.END, f"{[str(c)[:15] for c in r]}\n")
            except Exception as e: messagebox.showerror("错误", str(e))
        tk.Button(win, text="筛选并导出", command=_go, bg=self.colors['primary'], fg="white",
                 font=("微软雅黑",11,"bold"), cursor="hand2").pack(pady=5)

    def _stats_excel_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        win = tk.Toplevel(self.root); win.title("数据统计分析"); win.geometry("720x520")
        win.transient(self.root); win.grab_set()

        # 顶栏信息
        top_f = tk.Frame(win, bg=self.colors['light']); top_f.pack(fill=tk.X, padx=15, pady=8)
        tk.Label(top_f, text="📊 数据统计分析", font=("微软雅黑", 14, "bold"),
                 bg=self.colors['light']).pack(side=tk.LEFT)
        status_var = tk.StringVar(value="正在读取...")
        tk.Label(top_f, textvariable=status_var, font=("微软雅黑", 9),
                 fg="gray", bg=self.colors['light']).pack(side=tk.RIGHT)

        # Notebook: 统计结果 + 原始数据预览
        nb = ttk.Notebook(win); nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Tab1: 统计结果
        tab1 = tk.Frame(nb); nb.add(tab1, text="📈 统计结果")
        log = scrolledtext.ScrolledText(tab1, height=20, font=("Consolas", 10))
        log.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)

        # Tab2: 原始数据预览
        tab2 = tk.Frame(nb); nb.add(tab2, text="📋 数据预览")
        preview = scrolledtext.ScrolledText(tab2, height=20, font=("Consolas", 10))
        preview.pack(fill=tk.BOTH, expand=True, padx=8, pady=5)

        try:
            wb = openpyxl.load_workbook(fp, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"无法读取文件: {e}"); return

        if not all_rows:
            log.insert(tk.END, "文件为空\n"); return

        hdr = [str(c) for c in all_rows[0]]
        data_rows = all_rows[1:]
        total_rows = len(data_rows)
        total_cols = len(hdr)

        # 数据预览（前30行）
        preview.insert(tk.END, f"📄 {Path(fp).name}  |  共 {total_rows} 行 × {total_cols} 列\n\n")
        preview.insert(tk.END, f"{'行号':>4} │ " + " │ ".join(f"{h:<12}" for h in hdr[:6]) + "\n")
        preview.insert(tk.END, "─" * 80 + "\n")
        for i, r in enumerate(data_rows[:30], 1):
            vals = [str(v)[:12] if v is not None else "" for v in r[:6]]
            preview.insert(tk.END, f"{i:>4} │ " + " │ ".join(f"{v:<12}" for v in vals) + "\n")
        if total_rows > 30:
            preview.insert(tk.END, f"\n... 仅显示前30行，共 {total_rows} 行\n")

        # 统计分析
        status_var.set(f"分析 {total_cols} 列...")
        win.update()

        from collections import Counter
        import math

        log.insert(tk.END, f"{'='*60}\n")
        log.insert(tk.END, f"  万能办公助手 — 数据统计分析报告\n")
        log.insert(tk.END, f"  文件: {Path(fp).name}\n")
        log.insert(tk.END, f"  规模: {total_rows} 行 × {total_cols} 列\n")
        log.insert(tk.END, f"{'='*60}\n\n")

        numeric_cols = 0
        text_cols = 0

        for ci in range(total_cols):
            col_name = hdr[ci] if ci < len(hdr) else f"列{ci+1}"
            vals = []
            nulls = 0
            for r in data_rows:
                v = r[ci] if ci < len(r) else None
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    nulls += 1
                else:
                    vals.append(v)

            non_null = len(vals)
            # 判断是否数值列
            num_vals = []
            text_vals = []
            for v in vals:
                try:
                    num_vals.append(float(v))
                except (ValueError, TypeError):
                    text_vals.append(str(v))

            is_numeric = len(num_vals) >= non_null * 0.8 and non_null > 0

            if is_numeric:
                numeric_cols += 1
                n = len(num_vals)
                s = sorted(num_vals)
                mn, mx = s[0], s[-1]
                avg = sum(num_vals) / n
                var = sum((x - avg)**2 for x in num_vals) / n
                sd = var ** 0.5
                # 中位数
                med = s[n//2] if n % 2 else (s[n//2-1] + s[n//2]) / 2
                # 四分位数
                q1 = s[int(n * 0.25)]
                q3 = s[int(n * 0.75)]
                iqr = q3 - q1
                # 众数
                counter = Counter(num_vals)
                mode_val, mode_cnt = counter.most_common(1)[0]
                # 偏度/峰度
                skew = sum((x - avg)**3 for x in num_vals) / (n * sd**3) if sd > 0 else 0
                kurt = sum((x - avg)**4 for x in num_vals) / (n * var**2) - 3 if var > 0 else 0

                log.insert(tk.END, f"  📊 【{col_name}】 数值列  (n={n}, 空值={nulls})\n")
                log.insert(tk.END, f"  {'─'*50}\n")
                log.insert(tk.END, f"    集中趋势: 平均={avg:.4f}  中位数={med:.4f}  众数={mode_val:.4f}({mode_cnt}次)\n")
                log.insert(tk.END, f"    离散程度: 标准差={sd:.4f}  方差={var:.4f}\n")
                log.insert(tk.END, f"    极值:     最小值={mn:.4f}  最大值={mx:.4f}  极差={mx-mn:.4f}\n")
                log.insert(tk.END, f"    分位数:   25%={q1:.4f}  75%={q3:.4f}  IQR={iqr:.4f}\n")
                log.insert(tk.END, f"    分布形态: 偏度={skew:.4f}  峰度={kurt:.4f}\n")
                log.insert(tk.END, f"    总和={sum(num_vals):.2f}\n\n")
            else:
                text_cols += 1
                n = len(text_vals)
                counter = Counter(text_vals)
                top5 = counter.most_common(5)
                unique = len(counter)

                log.insert(tk.END, f"  📝 【{col_name}】 文本列  (总数={non_null}, 空值={nulls})\n")
                log.insert(tk.END, f"  {'─'*50}\n")
                log.insert(tk.END, f"    非空值: {non_null}  唯一值: {unique}  空值: {nulls}\n")
                log.insert(tk.END, f"    最长值: {max(len(str(v)) for v in text_vals)} 字符\n")
                log.insert(tk.END, f"    最短值: {min(len(str(v)) for v in text_vals)} 字符\n")
                log.insert(tk.END, f"    前5高频:\n")
                for val, cnt in top5:
                    pct = cnt / n * 100 if n > 0 else 0
                    bar = "█" * int(pct / 5)
                    log.insert(tk.END, f"      {bar} {val[:20]:<20} {cnt:>4}次 ({pct:.1f}%)\n")
                log.insert(tk.END, "\n")

        # 总体摘要
        log.insert(tk.END, f"{'='*60}\n")
        log.insert(tk.END, f"  📋 总体摘要\n")
        log.insert(tk.END, f"  总行数: {total_rows}  |  总列数: {total_cols}\n")
        log.insert(tk.END, f"  数值列: {numeric_cols}  |  文本列: {text_cols}\n")
        log.insert(tk.END, f"  文件名: {Path(fp).name}\n")
        log.insert(tk.END, f"  工作表: {ws.title}\n")
        log.insert(tk.END, f"{'='*60}\n")

        # 状态栏
        status_var.set(f"✅ 分析完成: {numeric_cols}数值列 + {text_cols}文本列")
        self.set_status(f"统计分析: {Path(fp).name}")

        # 保存按钮
        btn_f = tk.Frame(win, bg=win.cget('bg')); btn_f.pack(pady=5)
        tk.Button(btn_f, text="💾 导出报告", command=lambda: _save_report(), cursor="hand2",
                  bg=self.colors['primary'], fg="white", font=("微软雅黑", 9)).pack()
        def _save_report():
            save = filedialog.asksaveasfilename(defaultextension=".txt",
                     filetypes=[("文本文件","*.txt")])
            if save:
                Path(save).write_text(log.get("1.0", tk.END), encoding="utf-8")
                self.set_status(f"报告已保存: {Path(save).name}")

    def _excel_to_csv_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        save = filedialog.asksaveasfilename(title="保存CSV", defaultextension=".csv", filetypes=[("CSV","*.csv")])
        if not save: return
        def _go():
            try:
                wb = openpyxl.load_workbook(fp, read_only=True); ws = wb.active
                import csv
                with open(save, 'w', newline='', encoding='utf-8-sig') as f:
                    w = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        w.writerow([str(c) if c is not None else "" for c in row])
                wb.close()
                self.root.after(0, lambda: self._show_success_dialog("完成", f"导出CSV成功\n{save}"))
            except Exception as e: self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
        self._run_thread(_go, done_msg="导出完成")

    def _excel_chart_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        win = tk.Toplevel(self.root); win.title("创建图表"); win.geometry("400x280")
        win.transient(self.root); win.grab_set()
        wb = openpyxl.load_workbook(fp, read_only=True); all_rows = list(wb.active.iter_rows(values_only=True)); wb.close()
        hdr = [str(c) for c in all_rows[0]] if all_rows else []
        tk.Label(win, text="选择图表类型:", font=("微软雅黑",10)).pack(pady=10)
        ct = ttk.Combobox(win, values=["柱状图","折线图","饼图","条形图"], state="readonly", width=12); ct.set("柱状图"); ct.pack()
        tk.Label(win, text="分类列:", font=("微软雅黑",10)).pack(pady=5)
        c = ttk.Combobox(win, values=hdr, state="readonly", width=16)
        if hdr: c.set(hdr[0])
        c.pack()
        tk.Label(win, text="数值列:", font=("微软雅黑",10)).pack(pady=5)
        v = ttk.Combobox(win, values=hdr, state="readonly", width=16)
        if len(hdr)>1: v.set(hdr[1])
        v.pack()
        def _go():
            try:
                ci = hdr.index(c.get()); vi = hdr.index(v.get())
                data = []; cats = []
                for r in all_rows[1:]:
                    if r[ci] is not None and r[vi] is not None:
                        cats.append(str(r[ci])); data.append(float(r[vi]))
                wb2 = openpyxl.Workbook(); ws2 = wb2.active
                ws2.append([hdr[ci], hdr[vi]])
                for i in range(len(cats)): ws2.append([cats[i], data[i]])
                chart_type = ct.get()
                if "柱状" in chart_type:
                    from openpyxl.chart import BarChart as Chart
                    chart = Chart()
                elif "折线" in chart_type:
                    from openpyxl.chart import LineChart as Chart
                    chart = Chart()
                elif "饼图" in chart_type:
                    from openpyxl.chart import PieChart as Chart
                    chart = Chart()
                else:
                    from openpyxl.chart import BarChart as Chart
                    chart = Chart(); chart.type = "bar"
                from openpyxl.chart import Reference
                data_ref = Reference(ws2, min_col=2, min_row=1, max_row=len(cats)+1)
                cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(cats)+1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref); chart.title = f"{hdr[vi]} {chart_type}"
                ws2.add_chart(chart, "E5")
                save = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
                if save: wb2.save(save)
                wb2.close()
                if save: self.root.after(0, lambda: self._show_success_dialog("完成", f"图表创建成功\n{save}"))
            except Exception as e: messagebox.showerror("错误", str(e))
        tk.Button(win, text="生成图表", command=_go, bg=self.colors['primary'], fg="white",
                 font=("微软雅黑",11,"bold"), cursor="hand2", width=12).pack(pady=12)

    def _vlookup_dlg(self):
        files = filedialog.askopenfilenames(title="选择两个Excel（主表+查询表）", filetypes=[("Excel","*.xlsx")])
        if len(files) != 2: return messagebox.showwarning("提示", "请选择两个文件")
        win = tk.Toplevel(self.root); win.title("VLOOKUP"); win.geometry("500x300")
        win.transient(self.root); win.grab_set()
        wb1 = openpyxl.load_workbook(files[0], read_only=True); rows1 = list(wb1.active.iter_rows(values_only=True)); wb1.close()
        wb2 = openpyxl.load_workbook(files[1], read_only=True); rows2 = list(wb2.active.iter_rows(values_only=True)); wb2.close()
        hdr1 = [str(c) for c in rows1[0]]; hdr2 = [str(c) for c in rows2[0]]
        tk.Label(win, text="主表匹配列:", font=("微软雅黑",10)).pack()
        key = ttk.Combobox(win, values=hdr1, state="readonly", width=16); key.set(hdr1[0] if hdr1 else ""); key.pack()
        tk.Label(win, text="查询表匹配列:", font=("微软雅黑",10)).pack(pady=5)
        ex = ttk.Combobox(win, values=hdr2, state="readonly", width=16); ex.set(hdr2[0] if hdr2 else ""); ex.pack()
        def _go():
            try:
                ki = hdr1.index(key.get()); ei = hdr2.index(ex.get())
                key_map = {str(r[ei]): [str(c) for c in r] for r in rows2[1:] if r[ei] is not None}
                wb_out = openpyxl.Workbook(); ws_out = wb_out.active
                ws_out.append(hdr1 + [f"{Path(files[1]).stem}_{c}" for c in hdr2])
                for r in rows1[1:]:
                    match_val = str(r[ki]) if r[ki] is not None else ""
                    match_row = key_map.get(match_val, [""]*len(hdr2))
                    ws_out.append([str(c) if c is not None else "" for c in r] + match_row)
                save = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
                if save: wb_out.save(save)
                wb_out.close()
                if save: self.root.after(0, lambda: self._show_success_dialog("完成", f"VLOOKUP完成\n{save}"))
            except Exception as e: messagebox.showerror("错误", str(e))
        tk.Button(win, text="执行VLOOKUP", command=_go, bg=self.colors['primary'], fg="white",
                 font=("微软雅黑",11,"bold"), cursor="hand2", width=14).pack(pady=12)

    def _conditional_format_dlg(self):
        fp = filedialog.askopenfilename(title="选择Excel", filetypes=[("Excel","*.xlsx")])
        if not fp: return
        win = tk.Toplevel(self.root); win.title("条件格式"); win.geometry("500x350")
        win.transient(self.root); win.grab_set()
        wb = openpyxl.load_workbook(fp, read_only=True); all_rows = list(wb.active.iter_rows(values_only=True)); wb.close()
        hdr = [str(c) for c in all_rows[0]] if all_rows else []
        tk.Label(win, text="选择要格式化的列:", font=("微软雅黑",10)).pack(pady=8)
        lb = tk.Listbox(win, height=5, font=("微软雅黑",9)); lb.pack(fill=tk.X, padx=20)
        for h in hdr: lb.insert(tk.END, h); lb.selection_set(0)
        tk.Label(win, text="条件 (>0, <100, ==0, !=0):", font=("微软雅黑",10)).pack(pady=5)
        e = tk.Entry(win, font=("微软雅黑",10), width=20); e.insert(0, "<60"); e.pack()
        def _go():
            try:
                sel = lb.curselection(); ci = sel[0] if sel else 0
                cond = e.get().strip(); cond_v = cond[2:].strip() if cond[:2] in (">=","<=","==","!=") else cond[1:].strip()
                wb = openpyxl.load_workbook(fp); ws = wb.active
                fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                    val = row[ci].value
                    if val is not None:
                        try:
                            nv = float(val)
                            ex = cond
                            if ex.startswith(">="): ok = nv >= float(ex[2:])
                            elif ex.startswith("<="): ok = nv <= float(ex[2:])
                            elif ex.startswith("=="): ok = nv == float(ex[2:])
                            elif ex.startswith("!="): ok = nv != float(ex[2:])
                            elif ex.startswith(">"): ok = nv > float(ex[1:])
                            elif ex.startswith("<"): ok = nv < float(ex[1:])
                            else: ok = False
                            if ok: row[ci].fill = fill
                        except: pass
                save = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
                if save: wb.save(save)
                wb.close()
                if save: self.root.after(0, lambda: self._show_success_dialog("完成", f"条件格式完成\n{save}"))
            except Exception as e: messagebox.showerror("错误", str(e))
        tk.Button(win, text="应用", command=_go, bg=self.colors['primary'], fg="white",
                 font=("微软雅黑",11,"bold"), cursor="hand2", width=12).pack(pady=12)
